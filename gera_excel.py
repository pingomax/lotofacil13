# -*- coding: utf-8 -*-
"""Gera planilha profissional Analise_Lotofacil_300_Concursos_e_9_Jogos.xlsx"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

o = json.load(open("resultados.json"))
NUMS = list(range(1,26))
OUT = "Analise_Lotofacil_300_Concursos_e_9_Jogos.xlsx"

# paleta
AZUL = "1F3864"; AZUL2="2E5496"; CINZA="D9D9D9"; VERDE="C6EFCE"; VERM="FFC7CE"
LARANJA="FFC000"; AMARELO="FFEB9C"; AZULCLARO="BDD7EE"
hdr_fill = PatternFill("solid", fgColor=AZUL)
hdr_font = Font(color="FFFFFF", bold=True, size=11)
title_font = Font(color=AZUL, bold=True, size=14)
sub_font = Font(bold=True, size=11, color=AZUL2)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin,right=thin,top=thin,bottom=thin)

def style_header(ws, row, ncols, start=1):
    for c in range(start, start+ncols):
        cell = ws.cell(row=row, column=c)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = center; cell.border = border

def autosize(ws, widths):
    for col,w in widths.items():
        ws.column_dimensions[col].width = w

def title(ws, text, cell="A1"):
    ws[cell] = text; ws[cell].font = title_font

wb = Workbook()
wb.remove(wb.active)

# classe cores
CLS_FILL = {"Muito quente":PatternFill("solid",fgColor="C00000"),
            "Quente":PatternFill("solid",fgColor="FF8C00"),
            "Neutra":PatternFill("solid",fgColor="FFD966"),
            "Fria":PatternFill("solid",fgColor="9DC3E6"),
            "Muito fria":PatternFill("solid",fgColor="2E75B6")}
CLS_FONT = {"Muito quente":Font(color="FFFFFF",bold=True),"Quente":Font(bold=True),
            "Neutra":Font(bold=True),"Fria":Font(bold=True),"Muito fria":Font(color="FFFFFF",bold=True)}
# ===== ABA 2: BASE_300_CONCURSOS =====
ws = wb.create_sheet("Base_300_Concursos")
hdr = ["Concurso","Data"]+[f"D{i}" for i in range(1,16)]+["Soma","Pares","Impares","Moldura","Centro","Primos"]
ws.append(hdr); style_header(ws,1,len(hdr))
PRIMOS={2,3,5,7,11,13,17,19,23}; CENTRO={7,8,9,12,13,14,17,18,19}; PAR={n for n in NUMS if n%2==0}
for row in o['base']:
    conc,data = row[0],row[1]; dz=row[2:]
    s=set(dz); p=len(s&PAR); m=len(s-CENTRO)
    ws.append([conc,data]+dz+[sum(dz),p,15-p,m,15-m,len(s&PRIMOS)])
ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:{get_column_letter(len(hdr))}1"
autosize(ws,{"A":10,"B":12})
for c in range(3,18): ws.column_dimensions[get_column_letter(c)].width=5
for r in range(2,ws.max_row+1):
    for c in range(1,len(hdr)+1):
        ws.cell(r,c).alignment=center; ws.cell(r,c).border=border

# ===== ABA 3: FREQUENCIA_DEZENAS =====
ws = wb.create_sheet("Frequencia_Dezenas")
title(ws,"Frequência das Dezenas — 300 concursos mais recentes")
hdr=["Dezena","Aparições (300)","% (300)","Esperada","Dif. vs esperada","Freq 100","Freq 50","Freq 20","Freq 10","Atraso atual","Maior atraso","Seq. máx.","Classe"]
ws.append([]); ws.append(hdr); style_header(ws,3,len(hdr))
esp=o['freq']['esperada']
for n in NUMS:
    f300=o['freq']['f300'].get(str(n),0)
    ws.append([n,f300,round(f300/300*100,1),round(esp,1),round(f300-esp,1),
               o['freq']['f100'].get(str(n),0),o['freq']['f50'].get(str(n),0),
               o['freq']['f20'].get(str(n),0),o['freq']['f10'].get(str(n),0),
               o['freq']['atraso'][str(n)],o['freq']['maior_atraso'][str(n)],
               o['freq']['seq_max'][str(n)],o['freq']['classe'][str(n)]])
ws.freeze_panes="A4"
for r in range(4,ws.max_row+1):
    cl=ws.cell(r,13).value
    ws.cell(r,13).fill=CLS_FILL[cl]; ws.cell(r,13).font=CLS_FONT[cl]
    for c in range(1,len(hdr)+1): ws.cell(r,c).alignment=center; ws.cell(r,c).border=border
autosize(ws,{"A":8,"B":15,"C":9,"D":10,"E":15,"F":9,"G":8,"H":8,"I":8,"J":12,"K":12,"L":10,"M":13})
# grafico de barras freq300
chart=BarChart(); chart.title="Frequência nos 300 concursos"; chart.type="col"; chart.height=8; chart.width=20
data=Reference(ws,min_col=2,min_row=3,max_row=28); cats=Reference(ws,min_col=1,min_row=4,max_row=28)
chart.add_data(data,titles_from_data=True); chart.set_categories(cats); chart.legend=None
ws.add_chart(chart,"O3")

# ===== ABA 4: QUENTES_E_FRIAS =====
ws=wb.create_sheet("Quentes_e_Frias")
title(ws,"Classificação por temperatura (frequência + tendência + atraso)")
r=3
def bloco(ws,r,titulo,lista,mostra_classe=True):
    ws.cell(r,1,titulo).font=sub_font; r+=1
    ws.cell(r,1,"Dezena").fill=hdr_fill; ws.cell(r,1).font=hdr_font
    ws.cell(r,2,"Classe" if mostra_classe else "Valor").fill=hdr_fill; ws.cell(r,2).font=hdr_font
    ws.cell(r,1).alignment=center; ws.cell(r,2).alignment=center; r+=1
    for item in lista:
        if isinstance(item,tuple): n,v=item
        else: n,v=item,o['freq']['classe'].get(str(item),"")
        ws.cell(r,1,n).alignment=center; ws.cell(r,1).border=border
        ws.cell(r,2,v).alignment=center; ws.cell(r,2).border=border
        if mostra_classe and v in CLS_FILL: ws.cell(r,2).fill=CLS_FILL[v]; ws.cell(r,2).font=CLS_FONT[v]
        r+=1
    return r+1
ordp=o['freq']['ordenado']
r=bloco(ws,r,"10 mais QUENTES",ordp[:10])
r=bloco(ws,r,"5 NEUTRAS relevantes",ordp[10:15])
r=bloco(ws,r,"10 mais FRIAS",ordp[-10:])
r2=3
r2=bloco(ws,r2,"Tendência de ALTA",o['freq']['tend_alta'],False) if False else r2
# colunas laterais
ws.cell(3,4,"Tendência de ALTA").font=sub_font
for i,n in enumerate(o['freq']['tend_alta']): ws.cell(4+i,4,n).alignment=center; ws.cell(4+i,4).border=border
ws.cell(3,6,"Tendência de QUEDA").font=sub_font
for i,n in enumerate(o['freq']['tend_queda']): ws.cell(4+i,6,n).alignment=center; ws.cell(4+i,6).border=border
ws.cell(12,4,"Mais ATRASADAS").font=sub_font
for i,n in enumerate(o['freq']['atrasadas']): ws.cell(13+i,4,n).alignment=center; ws.cell(13+i,4).border=border
ws.cell(12,6,"Voltaram recentemente").font=sub_font
for i,n in enumerate(o['freq']['voltaram']): ws.cell(13+i,6,n).alignment=center; ws.cell(13+i,6).border=border
autosize(ws,{"A":20,"B":14,"D":20,"E":4,"F":20})
print("ABAS_2_4_OK")

def tabela(ws, start_row, headers, rows, col=1):
    for j,h in enumerate(headers):
        c=ws.cell(start_row,col+j,h); c.fill=hdr_fill; c.font=hdr_font; c.alignment=center; c.border=border
    for i,rowdata in enumerate(rows):
        for j,v in enumerate(rowdata):
            c=ws.cell(start_row+1+i,col+j,v); c.alignment=center; c.border=border
    return start_row+1+len(rows)+1

# ===== ABA 5: PARES_IMPARES =====
ws=wb.create_sheet("Pares_Impares")
title(ws,"Distribuição Pares x Ímpares — 300 concursos")
rows=[[f"{k[0]}P / {k[1]}I",v,f"{v/300*100:.1f}%"] for k,v in o['pi']]
nr=tabela(ws,3,["Padrão (Pares/Ímpares)","Ocorrências","%"],rows)
ws.cell(nr,1,"Últimos 10 concursos").font=sub_font
tabela(ws,nr+1,["Padrão","Ocorr."],[[f"{k[0]}P/{k[1]}I",v] for k,v in o['pi10']])
ws.cell(nr,5,"Últimos 20 concursos").font=sub_font
tabela(ws,nr+1,["Padrão","Ocorr."],[[f"{k[0]}P/{k[1]}I",v] for k,v in o['pi20']],col=5)
autosize(ws,{"A":22,"B":14,"C":8,"E":22,"F":10})
ch=BarChart(); ch.title="Pares/Ímpares"; ch.type="col"; ch.height=8; ch.width=14
d=Reference(ws,min_col=2,min_row=3,max_row=3+len(rows)); c=Reference(ws,min_col=1,min_row=4,max_row=3+len(rows))
ch.add_data(d,titles_from_data=True); ch.set_categories(c); ch.legend=None; ws.add_chart(ch,"H3")

# ===== ABA 6: MOLDURA_CENTRO =====
ws=wb.create_sheet("Moldura_Centro")
title(ws,"Moldura x Centro — 300 concursos")
rows=[[f"{k[0]}M / {k[1]}C",v,f"{v/300*100:.1f}%"] for k,v in o['mc']]
tabela(ws,3,["Padrão (Moldura/Centro)","Ocorrências","%"],rows)
ws.cell(3,5,"Padrão mais comum por janela").font=sub_font
jr=[["Janela","Padrão","Ocorr."]]
for lbl,key in [("10",'mc10'),("20",'mc20'),("50",'mc50'),("100",'mc100'),("300",'mc')]:
    top=o[key][0]; jr.append([f"Últ {lbl}",f"{top[0][0]}M/{top[0][1]}C",top[1]])
for i,rr in enumerate(jr):
    for j,v in enumerate(rr):
        cc=ws.cell(4+i,5+j,v); cc.alignment=center; cc.border=border
        if i==0: cc.fill=hdr_fill; cc.font=hdr_font
autosize(ws,{"A":22,"B":14,"C":8,"E":10,"F":12,"G":8})
ch=BarChart(); ch.title="Moldura/Centro"; ch.type="col"; ch.height=8; ch.width=14
d=Reference(ws,min_col=2,min_row=3,max_row=3+len(rows)); c=Reference(ws,min_col=1,min_row=4,max_row=3+len(rows))
ch.add_data(d,titles_from_data=True); ch.set_categories(c); ch.legend=None; ws.add_chart(ch,"I3")
print("ABAS_5_6_OK")

# ===== ABA 7: LINHAS =====
ws=wb.create_sheet("Linhas")
title(ws,"Análise das 5 Linhas (matriz 5x5)")
rows=[]
for l in range(1,6):
    tot=o['linha_totais'][str(l)]; media=tot/300
    ocup=o['linha_ocup'][str(l)]
    mn=min(int(k) for k in ocup); mx=max(int(k) for k in ocup)
    rows.append([f"Linha {l}",tot,round(media,2),mn,mx])
tabela(ws,3,["Linha","Freq. total","Média/conc.","Mín","Máx"],rows)
ws.cell(10,1,"Distribuição de ocupação (0..5 dezenas) por linha").font=sub_font
hdrs=["Linha"]+[f"{k} dez." for k in range(6)]
orows=[]
for l in range(1,6):
    ocup=o['linha_ocup'][str(l)]
    orows.append([f"Linha {l}"]+[ocup.get(str(k),0) for k in range(6)])
tabela(ws,11,hdrs,orows)
ws.cell(19,1,"Distribuições de linha mais frequentes (L1-L2-L3-L4-L5)").font=sub_font
drows=[["-".join(map(str,k)),v,f"{v/300*100:.1f}%"] for k,v in o['linha_dist'][:12]]
tabela(ws,20,["Distribuição","Ocorr.","%"],drows)
autosize(ws,{"A":18,"B":12,"C":12,"D":6,"E":6})

# ===== ABA 8: COLUNAS =====
ws=wb.create_sheet("Colunas")
title(ws,"Análise das 5 Colunas (matriz 5x5)")
rows=[]
for cc in range(1,6):
    tot=o['col_totais'][str(cc)]; media=tot/300
    ocup=o['col_ocup'][str(cc)]
    mn=min(int(k) for k in ocup); mx=max(int(k) for k in ocup)
    rows.append([f"Coluna {cc}",tot,round(media,2),mn,mx])
tabela(ws,3,["Coluna","Freq. total","Média/conc.","Mín","Máx"],rows)
ws.cell(10,1,"Distribuição de ocupação (0..5) por coluna").font=sub_font
orows=[]
for cc in range(1,6):
    ocup=o['col_ocup'][str(cc)]
    orows.append([f"Coluna {cc}"]+[ocup.get(str(k),0) for k in range(6)])
tabela(ws,11,["Coluna"]+[f"{k} dez." for k in range(6)],orows)
ws.cell(19,1,"Distribuições de coluna mais frequentes (C1-C2-C3-C4-C5)").font=sub_font
drows=[["-".join(map(str,k)),v,f"{v/300*100:.1f}%"] for k,v in o['col_dist'][:12]]
tabela(ws,20,["Distribuição","Ocorr.","%"],drows)
autosize(ws,{"A":18,"B":12,"C":12,"D":6,"E":6})

# ===== ABA 9: SEQUENCIAS =====
ws=wb.create_sheet("Sequencias")
title(ws,"Sequências de dezenas consecutivas")
rows=[["Média de blocos consec./concurso",round(o['seq_por_conc_media'],2)],
      ["Média de dezenas em sequência",round(o['seq_dezenas_media'],2)],
      ["Maior sequência observada",o['maior_seq']]]
tabela(ws,3,["Métrica","Valor"],rows)
ws.cell(8,1,"Frequência por tamanho de bloco").font=sub_font
labels={2:"Pares consecutivos (2)",3:"Trincas (3)",4:"Sequência de 4",5:"Sequência de 5+"}
srows=[]
for sz in sorted(int(k) for k in o['seq_size']):
    srows.append([labels.get(sz,f"Bloco de {sz}"),o['seq_size'][str(sz)]])
tabela(ws,9,["Tipo","Ocorrências (total de blocos)"],srows)
autosize(ws,{"A":32,"B":26})
print("ABAS_7_9_OK")

# ===== ABA 10: REPETICOES =====
ws=wb.create_sheet("Repeticoes")
title(ws,"Repetição de dezenas entre concursos consecutivos")
rp=o['rep']
rows=[["Média",round(rp['media'],2)],["Mínimo",rp['min']],["Máximo",rp['max']],["Moda",rp['moda']]]
tabela(ws,3,["Métrica","Valor"],rows)
ws.cell(9,1,"Distribuição da quantidade de repetidas").font=sub_font
crows=[]
for k in sorted(int(x) for x in rp['counter']):
    v=rp['counter'][str(k)]; crows.append([k,v,f"{v/300*100:.1f}%"])
tabela(ws,10,["Repetidas","Ocorrências","%"],crows)
ws.cell(9,6,"Permanência por dezena (apareceu -> repetiu)").font=sub_font
prows=[]
for n in NUMS:
    ap,re=rp['perm'][str(n)]; prows.append([n,ap,re,f"{(re/ap*100 if ap else 0):.0f}%"])
tabela(ws,10,["Dezena","Apareceu","Repetiu","% perman."],prows,col=6)
autosize(ws,{"A":14,"B":12,"C":8,"F":8,"G":10,"H":9,"I":10})

# ===== ABA 11: SOMA =====
ws=wb.create_sheet("Soma")
title(ws,"Soma das 15 dezenas")
sm=o['soma']
rows=[["Menor soma",sm['min']],["Maior soma",sm['max']],["Média",round(sm['media'],1)],["Mediana",sm['mediana']]]
tabela(ws,3,["Métrica","Valor"],rows)
ws.cell(9,1,"Distribuição por faixa de soma").font=sub_font
frows=[[k,v,f"{v/300*100:.1f}%"] for k,v in sm['faixas'].items()]
nr=tabela(ws,10,["Faixa","Ocorrências","%"],frows)
ch=BarChart(); ch.title="Soma por faixa"; ch.type="col"; ch.height=8; ch.width=14
d=Reference(ws,min_col=2,min_row=10,max_row=10+len(frows)); c=Reference(ws,min_col=1,min_row=11,max_row=10+len(frows))
ch.add_data(d,titles_from_data=True); ch.set_categories(c); ch.legend=None; ws.add_chart(ch,"F9")
autosize(ws,{"A":16,"B":14,"C":8})

# ===== ABA 12: BAIXAS_ALTAS =====
ws=wb.create_sheet("Baixas_Altas")
title(ws,"Dezenas Baixas (01-13) x Altas (14-25)")
rows=[[f"{k[0]}B / {k[1]}A",v,f"{v/300*100:.1f}%"] for k,v in o['ba']]
tabela(ws,3,["Padrão (Baixas/Altas)","Ocorrências","%"],rows)
autosize(ws,{"A":22,"B":14,"C":8})
ch=BarChart(); ch.title="Baixas/Altas"; ch.type="col"; ch.height=8; ch.width=14
d=Reference(ws,min_col=2,min_row=3,max_row=3+len(rows)); c=Reference(ws,min_col=1,min_row=4,max_row=3+len(rows))
ch.add_data(d,titles_from_data=True); ch.set_categories(c); ch.legend=None; ws.add_chart(ch,"E3")

# ===== ABA 13: PRIMOS_MULTIPLOS =====
ws=wb.create_sheet("Primos_Multiplos")
title(ws,"Números primos e múltiplos")
ws.cell(3,1,"Distribuição de primos por concurso (primos: 2,3,5,7,11,13,17,19,23)").font=sub_font
prows=[[k,v,f"{v/300*100:.1f}%"] for k,v in sorted(o['primo']['counter'].items(),key=lambda x:int(x[0]))]
tabela(ws,4,["Qtd. primos","Ocorrências","%"],prows)
ws.cell(3,6,f"Média de primos/concurso: {o['primo']['media']:.2f}").font=sub_font
ws.cell(15,1,"Média de múltiplos por concurso").font=sub_font
mrows=[["Múltiplos de 3",round(o['mult']['m3'],2)],["Múltiplos de 4",round(o['mult']['m4'],2)],["Múltiplos de 5",round(o['mult']['m5'],2)]]
tabela(ws,16,["Tipo","Média/conc."],mrows)
ws.cell(15,6,"Finais das dezenas (frequência total)").font=sub_font
frows=[[f"Final {k}",o['finais'].get(str(k),0)] for k in range(10)]
tabela(ws,16,["Final","Frequência"],frows,col=6)
autosize(ws,{"A":16,"B":14,"C":8,"F":12,"G":12})
print("ABAS_10_13_OK")

# ===== ABA 14: PARES_DE_DEZENAS =====
ws=wb.create_sheet("Pares_de_Dezenas")
title(ws,"Correlações — pares de dezenas que mais/menos saíram juntos")
top=[[f"{k[0]:02d}-{k[1]:02d}",v] for k,v in o['pares_top']]
tabela(ws,3,["Par","Ocorrências (300)"],top)
ws.cell(3,4,"20 pares menos frequentes").font=sub_font
bot=[[f"{k[0]:02d}-{k[1]:02d}",v] for k,v in o['pares_bot']]
tabela(ws,4,["Par","Ocorr."],bot,col=4)
ws.cell(3,7,"Trincas mais frequentes").font=sub_font
tri=[[f"{k[0]:02d}-{k[1]:02d}-{k[2]:02d}",v] for k,v in o['trincas_top']]
tabela(ws,4,["Trinca","Ocorr."],tri,col=7)
ws.cell(25,1,"Pares mais freq. — últimos 50 / 20 / 10").font=sub_font
for ci,(lbl,key) in enumerate([("50",'pares_50'),("20",'pares_20'),("10",'pares_10')]):
    prows=[[f"{k[0]:02d}-{k[1]:02d}",v] for k,v in o[key][:10]]
    tabela(ws,26,[f"Par (últ {lbl})","Ocorr."],prows,col=1+ci*3)
autosize(ws,{"A":14,"B":16,"D":10,"E":8,"G":14,"H":8,"J":6,"K":8})

# ===== ABA 15: MAPAS_DE_CALOR =====
ws=wb.create_sheet("Mapas_de_Calor")
title(ws,"Mapas de calor 5x5 — valores = frequência da dezena na janela")
janelas=[("Últimos 300",'f300'),("Últimos 100",'f100'),("Últimos 50",'f50'),("Últimos 20",'f20'),("Últimos 10",'f10')]
row0=3
for ji,(lbl,key) in enumerate(janelas):
    base_r = row0 + (ji//2)*8
    base_c = 1 + (ji%2)*7
    ws.cell(base_r,base_c,lbl).font=sub_font
    fdata=o['freq'][key]
    vals=[fdata.get(str(n),0) for n in NUMS]
    mn,mx=min(vals),max(vals)
    for i in range(5):
        for j in range(5):
            n=i*5+j+1; v=fdata.get(str(n),0)
            c=ws.cell(base_r+1+i,base_c+j,v); c.alignment=center; c.border=border
            # cor manual escala azul->vermelho
            t=(v-mn)/(mx-mn+1e-9)
            if t>0.8: c.fill=PatternFill("solid",fgColor="C00000"); c.font=Font(color="FFFFFF",bold=True)
            elif t>0.6: c.fill=PatternFill("solid",fgColor="FF8C00")
            elif t>0.4: c.fill=PatternFill("solid",fgColor="FFD966")
            elif t>0.2: c.fill=PatternFill("solid",fgColor="9DC3E6")
            else: c.fill=PatternFill("solid",fgColor="2E75B6"); c.font=Font(color="FFFFFF")
for col in "ABCDEFGHIJKLM": ws.column_dimensions[col].width=6
ws.cell(28,1,"Legenda: vermelho=muito quente, laranja=quente, amarelo=neutra, azul claro=fria, azul escuro=muito fria").font=Font(italic=True,size=9)
print("ABAS_14_15_OK")

# ===== ABA 16: ULTIMO_CONCURSO =====
ws=wb.create_sheet("Ultimo_Concurso")
u=o['ult']
title(ws,f"Concurso {u['concurso']} — {u['data']}")
ws.cell(3,1,"Dezenas sorteadas:").font=sub_font
for i,n in enumerate(u['dezenas']):
    c=ws.cell(4,1+i,n); c.alignment=center; c.border=border
    cl=o['freq']['classe'][str(n)]; c.fill=CLS_FILL[cl]; c.font=CLS_FONT[cl]
rows=[["Pares",u['pares']],["Ímpares",u['impares']],["Moldura",u['moldura']],["Centro",u['centro']],
      ["Baixas",u['baixas']],["Altas",u['altas']],["Soma",u['soma']],["Primos",u['primos']],
      ["Repetidas do anterior",u['repetidas_anterior']],
      ["Linhas (L1-L5)","-".join(map(str,u['linhas']))],
      ["Colunas (C1-C5)","-".join(map(str,u['colunas']))],
      ["Blocos consecutivos",", ".join(map(str,u['seqs'])) if u['seqs'] else "nenhum"]]
tabela(ws,6,["Atributo","Valor"],rows)
ws.cell(6,5,"Classe de cada dezena do último concurso").font=sub_font
crows=[[n,o['freq']['classe'][str(n)]] for n in u['dezenas']]
sr=8
for j,h in enumerate(["Dezena","Classe"]):
    cc=ws.cell(7,5+j,h); cc.fill=hdr_fill; cc.font=hdr_font; cc.alignment=center
for i,(n,cl) in enumerate(crows):
    ws.cell(sr+i,5,n).alignment=center; ws.cell(sr+i,5).border=border
    c=ws.cell(sr+i,6,cl); c.alignment=center; c.border=border; c.fill=CLS_FILL[cl]; c.font=CLS_FONT[cl]
# repeticao historica
perm=o['rep']['perm']
rep_ord=sorted(u['dezenas'],key=lambda n:(perm[str(n)][1]/perm[str(n)][0] if perm[str(n)][0] else 0),reverse=True)
ws.cell(22,1,"Maior taxa histórica de repetição (dentro do último concurso):").font=sub_font
ws.cell(23,1,"  "+", ".join(f"{n:02d} ({perm[str(n)][1]/perm[str(n)][0]*100:.0f}%)" for n in rep_ord[:5]))
ws.cell(24,1,"Menor taxa histórica de repetição:").font=sub_font
ws.cell(25,1,"  "+", ".join(f"{n:02d} ({perm[str(n)][1]/perm[str(n)][0]*100:.0f}%)" for n in rep_ord[-5:]))
ws.cell(26,1,"Obs.: taxas são estatística histórica, não previsão garantida.").font=Font(italic=True,size=9)
for col in "ABCDEFGHIJKLMNO": ws.column_dimensions[col].width=7
ws.column_dimensions["A"].width=22
print("ABA_16_OK")

# ===== ABA 17: JOGOS_SUGERIDOS =====
ws=wb.create_sheet("Jogos_Sugeridos")
title(ws,"9 Jogos sugeridos — critérios de equilíbrio estatístico")
hdr=["Jogo","Critério"]+[f"D{i}" for i in range(1,16)]+["Pares","Ímpares","Moldura","Centro","Rep.últ","Soma","Baixas","Altas","Primos","Seqs","Linhas","Colunas","Quentes","Neutras","Frias","Índice Eq. Estat."]
ws.append([]); ws.append(hdr); style_header(ws,3,len(hdr))
ult_set=set(o['ult']['dezenas'])
for i,g in enumerate(o['jogos']):
    m=o['metrics'][i]
    linhas="-".join(map(str,m['linhas'])); cols="-".join(map(str,m['colunas']))
    row=[f"Jogo {i+1}",o['criterios'][i]]+list(g)+[m['pares'],m['impares'],m['moldura'],m['centro'],
        m['repetidas_ult'],m['soma'],m['baixas'],m['altas'],m['primos'],m['seqs'],linhas,cols,
        m['quentes'],m['neutras'],m['frias'],o['idx_eq'][i]]
    ws.append(row)
ws.freeze_panes="C4"
# formatacao das dezenas por temperatura + marca repetidas
for i,g in enumerate(o['jogos']):
    r=4+i
    for j,n in enumerate(g):
        c=ws.cell(r,3+j)
        cl=o['freq']['classe'][str(n)]
        c.fill=CLS_FILL[cl]; c.font=CLS_FONT[cl]; c.alignment=center; c.border=border
        if n in ult_set:
            c.border=Border(left=Side(style="thick",color="000000"),right=Side(style="thick",color="000000"),
                            top=Side(style="thick",color="000000"),bottom=Side(style="thick",color="000000"))
    for c in range(1,len(hdr)+1):
        cell=ws.cell(r,c)
        if c>17: cell.alignment=center; cell.border=border
        if c<=2: cell.alignment=left; cell.border=border
ws.auto_filter.ref=f"A3:{get_column_letter(len(hdr))}3"
autosize(ws,{"A":8,"B":34})
for c in range(3,18): ws.column_dimensions[get_column_letter(c)].width=4.5
for c in range(18,len(hdr)+1): ws.column_dimensions[get_column_letter(c)].width=8
ws.column_dimensions[get_column_letter(len(hdr))].width=15
# legenda
lr=15
ws.cell(lr,1,"Legenda de cores:").font=sub_font
for k,(cl,txt) in enumerate([("Muito quente","Muito quente"),("Quente","Quente"),("Neutra","Neutra"),("Fria","Fria"),("Muito fria","Muito fria")]):
    c=ws.cell(lr+1+k,1,txt); c.fill=CLS_FILL[cl]; c.font=CLS_FONT[cl]; c.alignment=center; c.border=border
ws.cell(lr+7,1,"Borda preta grossa = dezena repetida do último concurso.").font=Font(italic=True,size=9)
print("ABA_17_OK")

# ===== ABA 18: CONFERENCIA (com formulas) =====
ws=wb.create_sheet("Conferencia")
title(ws,"Conferência — insira o resultado de um concurso")
ws.cell(3,1,"Resultado do concurso (digite as 15 dezenas sorteadas):").font=sub_font
for j in range(15):
    c=ws.cell(4,2+j); c.border=Border(left=Side(style="medium"),right=Side(style="medium"),top=Side(style="medium"),bottom=Side(style="medium"))
    c.fill=PatternFill("solid",fgColor="FFF2CC"); c.alignment=center
ws.cell(4,1,"Dezenas:").font=Font(bold=True)
# area de resultado das dezenas: B4:P4
res_range="$B$4:$P$4"
# tabela de jogos com acertos calculados por formula
hr=7
hdr=["Jogo"]+[f"D{i}" for i in range(1,16)]+["Acertos","Faixa premiação"]
for j,h in enumerate(hdr):
    c=ws.cell(hr,1+j,h); c.fill=hdr_fill; c.font=hdr_font; c.alignment=center; c.border=border
for i,g in enumerate(o['jogos']):
    r=hr+1+i
    ws.cell(r,1,f"Jogo {i+1}").alignment=center; ws.cell(r,1).border=border
    dez_cells=[]
    for j,n in enumerate(g):
        cell=ws.cell(r,2+j,n); cell.alignment=center; cell.border=border
        addr=f"{get_column_letter(2+j)}{r}"; dez_cells.append(addr)
        # verde se acertou
        cell.value=n
    # coluna acertos: soma de COUNTIF de cada dezena no resultado
    ac_col=17
    countifs="+".join(f"COUNTIF({res_range},{get_column_letter(2+j)}{r})" for j in range(15))
    ac_cell=ws.cell(r,ac_col); ac_cell.value=f"={countifs}"; ac_cell.alignment=center; ac_cell.border=border; ac_cell.font=Font(bold=True)
    fx_cell=ws.cell(r,ac_col+1)
    ac_addr=f"{get_column_letter(ac_col)}{r}"
    fx_cell.value=(f'=IF({ac_addr}=15,"15 pontos",IF({ac_addr}=14,"14 pontos",'
                   f'IF({ac_addr}=13,"13 pontos",IF({ac_addr}=12,"12 pontos",'
                   f'IF({ac_addr}=11,"11 pontos","Sem premiação")))))')
    fx_cell.alignment=center; fx_cell.border=border
# formatacao condicional: dezena que consta no resultado -> verde; senao -> vermelho
from openpyxl.formatting.rule import FormulaRule
green=PatternFill("solid",fgColor=VERDE); red=PatternFill("solid",fgColor=VERM)
for i in range(9):
    r=hr+1+i
    rng=f"B{r}:P{r}"
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f"COUNTIF({res_range},B{r})>0"],fill=green))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f"AND(B{r}<>\"\",COUNTIF({res_range},B{r})=0)"],fill=red))
# jogos premiados (>=11) em verde-claro na coluna faixa; melhor jogo destacado
ac_letter=get_column_letter(17)
prem_rng=f"R{hr+1}:R{hr+9}"
ws.conditional_formatting.add(prem_rng, FormulaRule(formula=[f'$Q{hr+1}>=11'],fill=PatternFill("solid",fgColor="E2EFDA")))
ac_rng=f"Q{hr+1}:Q{hr+9}"
ws.conditional_formatting.add(ac_rng, CellIsRule(operator="greaterThanOrEqual",formula=["11"],fill=green,font=Font(bold=True)))
# melhor jogo / resumo
sr=hr+11
ac_col_letter=get_column_letter(17)
first=hr+1; last=hr+9
ws.cell(sr,1,"Maior quantidade de acertos:").font=sub_font
ws.cell(sr,3).value=f"=MAX({ac_col_letter}{first}:{ac_col_letter}{last})"; ws.cell(sr,3).font=Font(bold=True)
ws.cell(sr+1,1,"Melhor jogo:").font=sub_font
ws.cell(sr+1,3).value=(f'="Jogo "&MATCH(MAX({ac_col_letter}{first}:{ac_col_letter}{last}),'
                       f'{ac_col_letter}{first}:{ac_col_letter}{last},0)')
ws.cell(sr+2,1,"Jogos premiados (>=11 acertos):").font=sub_font
ws.cell(sr+2,3).value=f"=COUNTIF({ac_col_letter}{first}:{ac_col_letter}{last},\">=11\")"
autosize(ws,{"A":30});
for c in range(2,17): ws.column_dimensions[get_column_letter(c)].width=5
ws.column_dimensions["Q"].width=8; ws.column_dimensions["R"].width=16
print("ABA_18_OK")

# ===== ABA 1: RESUMO (inserida no inicio) =====
ws=wb.create_sheet("Resumo",0)
title(ws,"Resumo Executivo — Análise Lotofácil (300 concursos)")
v=o['val']; u=o['ult']
def kv(r,k,val,bold=True):
    ws.cell(r,1,k).font=Font(bold=bold,color=AZUL2)
    ws.cell(r,3,val)
r=3
kv(r,"Período analisado",f"{v['data_ini']} a {v['data_fim']}"); r+=1
kv(r,"Concursos usados",f"{v['usados']} (de {v['total']} no arquivo)"); r+=1
kv(r,"Intervalo de concursos",f"{v['amostra_ini']} a {v['amostra_fim']}"); r+=1
kv(r,"Concurso mais recente",f"{u['concurso']} ({u['data']})"); r+=1
ws.cell(r,1,"Dezenas do último concurso").font=Font(bold=True,color=AZUL2)
ws.cell(r,3," ".join(f"{n:02d}" for n in u['dezenas'])); r+=2
ordp=o['freq']['ordenado']
kv(r,"10 dezenas mais QUENTES"," ".join(f"{n:02d}" for n in ordp[:10])); r+=1
kv(r,"5 dezenas NEUTRAS"," ".join(f"{n:02d}" for n in ordp[10:15])); r+=1
kv(r,"10 dezenas mais FRIAS"," ".join(f"{n:02d}" for n in ordp[-10:])); r+=1
kv(r,"Padrão pares/ímpares + comum",f"{o['pi'][0][0][0]}P / {o['pi'][0][0][1]}I ({o['pi'][0][1]}x)"); r+=1
kv(r,"Padrão moldura/centro + comum",f"{o['mc'][0][0][0]}M / {o['mc'][0][0][1]}C ({o['mc'][0][1]}x)"); r+=1
kv(r,"Qtd. repetidas + comum",f"{o['rep']['moda']} dezenas (média {o['rep']['media']:.1f})"); r+=1
faixa_top=max(o['soma']['faixas'].items(),key=lambda x:x[1])
kv(r,"Faixa de soma + comum",f"{faixa_top[0]} ({faixa_top[1]}x)"); r+=1
kv(r,"Distribuição de linhas + comum","-".join(map(str,o['linha_dist'][0][0]))); r+=1
kv(r,"Distribuição de colunas + comum","-".join(map(str,o['col_dist'][0][0]))); r+=1
kv(r,"Média de blocos consecutivos",f"{o['seq_por_conc_media']:.2f}"); r+=1
kv(r,"Tendência de ALTA"," ".join(f"{n:02d}" for n in o['freq']['tend_alta'])); r+=1
kv(r,"Tendência de QUEDA"," ".join(f"{n:02d}" for n in o['freq']['tend_queda'])); r+=1
kv(r,"Mais atrasadas"," ".join(f"{n:02d}" for n in o['freq']['atrasadas'])); r+=2
ws.cell(r,1,"Resumo dos 9 jogos").font=sub_font; r+=1
for j,h in enumerate(["Jogo","Dezenas","Índice Eq."]):
    c=ws.cell(r,1+ (0 if j==0 else (1 if j==1 else 6)),h); c.fill=hdr_fill; c.font=hdr_font; c.alignment=center
r+=1
for i,g in enumerate(o['jogos']):
    ws.cell(r,1,f"Jogo {i+1}").alignment=center
    ws.cell(r,2," ".join(f"{n:02d}" for n in g))
    ws.cell(r,7,o['idx_eq'][i]).alignment=center
    r+=1
r+=1
bt=o['bt']
ws.cell(r,1,"Teste retrospectivo (30 concursos, sem vazamento)").font=sub_font; r+=1
kv(r,"Média / Mediana de acertos",f"{bt['media']:.1f} / {bt['mediana']:.0f}"); r+=1
kv(r,"Faixa (mín-máx)",f"{bt['min']} a {bt['max']}"); r+=1
kv(r,"Jogos >=11 / >=12 / >=13",f"{bt['p11']} / {bt['p12']} / {bt['p13']}"); r+=2
ws.cell(r,1,"AVISO: Sorteios da Lotofácil são eventos aleatórios e independentes. Esta análise usa apenas").font=Font(italic=True,color="C00000",bold=True); r+=1
ws.cell(r,1,"padrões históricos e critérios de equilíbrio. Ela NÃO prevê resultados e NÃO altera a").font=Font(italic=True,color="C00000",bold=True); r+=1
ws.cell(r,1,"probabilidade matemática de uma aposta. Não há garantia de ganho.").font=Font(italic=True,color="C00000",bold=True)
autosize(ws,{"A":34,"B":40,"C":30,"G":12})
wb.save(OUT)
print("SALVO", OUT)
